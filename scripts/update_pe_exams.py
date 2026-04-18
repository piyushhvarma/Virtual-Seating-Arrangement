import os
import json
import re
import openpyxl
from copy import deepcopy

# Configuration
EXCEL_PATH = os.path.join("data", "pdfs", "ete", "PE-4,5,6  6th sem SEATING Plan.xlsx")
JSON_PATH = os.path.join("src", "data", "students.json")

# Map of extracted acronyms to proper IDs
SUBJECT_MAP = {
    'CV': ('AIM3240', 'Computer Vision'),
    'NLP': ('AIM3241', 'Natural Language Processing'),
    'SAOM': ('AIM3242', 'Sentiment Analysis and Opinion Mining'),
    'AICS': ('AIM3243', 'AI in Cyber Security'),
    'EAI': ('AIM3244', 'Explainable AI'),
    'GAI': ('AIM3245', 'Generative AI')
}

def get_default_subject(sheet_name):
    for key, val in SUBJECT_MAP.items():
        if key.upper() in sheet_name.upper():
            return val
    return ("UNKNOWN", "Unknown Subject")

def parse_excel():
    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        student_data = json.load(f)
        
    students = student_data["students"]
    
    # We want to clear out any old PE exams we inserted before if we run this script twice.
    # To do this safely, we keep ONLY exams that don't match the PE codes (AIM3240-3245)
    for reg, data in students.items():
        filtered_exams = []
        for e in data.get("exams", []):
            if not e["subjectCode"].startswith("AIM324"):
                filtered_exams.append(e)
        data["exams"] = filtered_exams

    total_added = 0
    missing_students = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        default_sub_code, default_sub_name = get_default_subject(sheet_name)
        
        # 1. Find Room Name
        room_name = "Unknown Room"
        for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and 'ROOM NO' in cell.upper():
                    match = re.search(r'ROOM NO:\s*([A-Za-z0-9\-\s\(\)]+)', cell, re.IGNORECASE)
                    if match:
                        # Clean up "LHC-003 Subject" string
                        r_name = match.group(1).split(' Subject')[0].strip()
                        r_name = r_name.split(' ROOM')[0].strip() # just in case
                        if r_name: room_name = r_name
        
        # 2. Build Legend
        legend = {}
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
            for cell in row:
                val = str(cell.value).strip() if cell.value else ''
                color = cell.fill.start_color.index if cell.fill.start_color else None
                if not color or str(color) == '00000000': continue
                
                if re.search(r'(?i)sec(tion)?', val) and len(val) < 50:
                    if 'ROOM NO' in val: continue # Ignore the huge header colored sections
                    legend[color] = val

        # 3. Find Data Grid Bounds
        start_row = -1
        end_row = -1
        cols_count = 0
        
        # find where 23FE begins
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True)):
            is_data_row = False
            for c_idx, cell in enumerate(row):
                val = str(cell).strip().upper() if cell else ""
                if val.startswith("23FE"):
                    is_data_row = True
                    cols_count = max(cols_count, c_idx + 1)
            
            if is_data_row and start_row == -1:
                start_row = r_idx + 1 # 1-indexed for openpyxl
            if is_data_row:
                end_row = r_idx + 1

        if start_row == -1:
            print(f"Skipping {sheet_name}, no student data found.")
            continue
            
        rows_count = (end_row - start_row) + 1
        
        # We want cols_count to just be the number of distinct columns containing data,
        # but to be visually accurate, we just assume the grid spans continuously.
        # Actually, let's find the specific columns that have headers like "COL A1".
        
        # Let's map column visual index -> logical column index
        # Some columns might be empty space. Let's find valid columns by looking at the row *above* start_row
        col_header_row = start_row - 1
        valid_cols = []
        for c_idx, cell in enumerate(ws[col_header_row]):
            val = str(cell.value).strip() if cell.value else ""
            if "COL" in val.upper() or val.startswith("23FE"): # just in case there's no header
                valid_cols.append(c_idx)
        
        if not valid_cols:
            # Fallback if no COL A1 headers
            for c_idx in range(cols_count): valid_cols.append(c_idx)

        # 4. Extract Students
        students_in_room = 0
        for excel_r in range(start_row, end_row + 1):
            for logical_c, excel_c in enumerate(valid_cols):
                cell = ws.cell(row=excel_r, column=excel_c + 1)
                val = str(cell.value).strip().upper() if cell.value else ""
                
                if val.startswith("23FE"):
                    students_in_room += 1
                    color = cell.fill.start_color.index if cell.fill.start_color else None
                    
                    # Resolve Subject/Section
                    legend_str = legend.get(color, "")
                    
                    sub_code = default_sub_code
                    sub_name = default_sub_name
                    section_assigned = "Unknown"
                    
                    if legend_str:
                        # Extract section
                        sec_match = re.search(r'(?i)sec(?:tion)?[\s\-.]*([A-Za-z]+)', legend_str)
                        if sec_match: section_assigned = sec_match.group(1).upper()
                        
                        # Extract specific subject if present in parenthesis
                        subj_match = re.search(r'\(([A-Za-z]+)\)', legend_str)
                        if subj_match:
                            acronym = subj_match.group(1).upper()
                            if acronym in SUBJECT_MAP:
                                sub_code, sub_name = SUBJECT_MAP[acronym]
                    
                    # Calculate index mapping
                    # col_idx = logical_c, row_idx = excel_r - start_row
                    row_idx = excel_r - start_row
                    seat_index = (logical_c * rows_count) + row_idx
                    
                    exam_record = {
                        "name": students.get(val, {}).get("name", "Unknown Name"),
                        "section": section_assigned,
                        "subjectCode": sub_code,
                        "subject": sub_name,
                        "room": room_name,
                        "seatIndex": seat_index,
                        "totalStudentsInRoom": 0, # Will update after loop
                        "rows": rows_count,
                        "cols": len(valid_cols),
                        "examDate": "TBD-APR-2026",
                        "examTime": "TBD"
                    }
                    
                    if val not in students:
                        missing_students += 1
                        students[val] = {
                            "name": "Unknown Output",
                            "exams": []
                        }
                    
                    students[val]["exams"].append(exam_record)
                    total_added += 1

        # Post Process: Update totalStudentsInRoom
        # Need to go back and update the dict elements we just added:
        for reg in students:
            for exam in students[reg]["exams"]:
                if exam["room"] == room_name and exam["subjectCode"].startswith("AIM324"):
                    exam["totalStudentsInRoom"] = students_in_room

    # Save
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(student_data, f, indent=4)

    print(f"Successfully processed PE exams.")
    print(f"Total PE Exam Tickets Issued: {total_added}")
    print(f"Total Missing Students (Added generically): {missing_students}")

if __name__ == "__main__":
    parse_excel()
